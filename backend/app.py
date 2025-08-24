from flask import render_template
from openpyxl import Workbook
from flask import send_file
import tempfile
import os
from qr_generator import generate_qr_code
from flask import Flask, jsonify, request
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from models import db, Tool, User, Project, ToolHistory

# Utworzenie aplikacji Flask
app = Flask(__name__)

# Konfiguracja
app.config['SECRET_KEY'] = 'tajny-klucz-zmien-to-pozniej'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JWT_SECRET_KEY'] = 'jwt-sekret-klucz'
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=24)

# Inicjalizacja rozszerzeń
db.init_app(app)
CORS(app, origins="*", supports_credentials=True)
jwt = JWTManager(app)

# Tworzenie tabel przy pierwszym uruchomieniu
with app.app_context():
    db.create_all()
    
    # Dodaj domyślnego admina jeśli nie istnieje
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(
            username='admin',
            email='admin@example.com',
            password_hash=generate_password_hash('admin123'),
            role='admin'
        )
        db.session.add(admin)
        db.session.commit()
        print("✓ Utworzono konto admina: admin / admin123")

# === STRONA GŁÓWNA ===
@app.route('/')
def index():
    """Strona główna - aplikacja webowa"""
    return render_template('index.html')

@app.route('/api')
def api_info():
    """Informacje o API"""
    return jsonify({
        'message': 'System Zarządzania Narzędziami - API',
        'endpoints': {
            'login': '/api/login',
            'tools': '/api/tools',
            'projects': '/api/projects'
        }
    })

# === LOGOWANIE ===
@app.route('/api/login', methods=['POST'])
def login():
    """Endpoint do logowania"""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    print(f"Próba logowania: {username}")
    
    user = User.query.filter_by(username=username).first()
    
    if user and check_password_hash(user.password_hash, password):
        access_token = create_access_token(identity=user.id)
        print(f"✓ Zalogowano: {username}")
        return jsonify({
            'access_token': access_token,
            'user': {
                'id': user.id,
                'username': user.username,
                'role': user.role
            }
        }), 200
    
    print(f"✗ Nieprawidłowe dane logowania dla: {username}")
    return jsonify({'message': 'Nieprawidłowa nazwa użytkownika lub hasło'}), 401

# === NARZĘDZIA ===
@app.route('/api/tools', methods=['GET'])
def get_tools():
    """Pobiera listę wszystkich narzędzi z informacją o wypożyczeniu"""
    tools = Tool.query.all()
    tools_data = []
    
    for tool in tools:
        tool_dict = tool.to_dict()
        
        # Jeśli narzędzie jest wypożyczone, dodaj info kto i na jaki projekt
        if tool.status == 'borrowed':
            # Znajdź ostatnie aktywne wypożyczenie
            last_borrow = ToolHistory.query.filter_by(
                tool_id=tool.id,
                returned_at=None
            ).order_by(ToolHistory.borrowed_at.desc()).first()
            
            if last_borrow:
                # Pobierz dane użytkownika
                user = User.query.get(last_borrow.user_id)
                tool_dict['borrowed_by'] = user.username if user else 'Nieznany'
                tool_dict['borrowed_by_id'] = last_borrow.user_id
                
                # Pobierz dane projektu
                if last_borrow.project_id:
                    project = Project.query.get(last_borrow.project_id)
                    tool_dict['project_name'] = project.name if project else 'Bez projektu'
                else:
                    tool_dict['project_name'] = 'Bez projektu'
                
                # Data wypożyczenia
                tool_dict['borrowed_date'] = last_borrow.borrowed_at.strftime('%Y-%m-%d %H:%M') if last_borrow.borrowed_at else ''
                
                print(f"Debug: {tool.name} wypożyczone przez {tool_dict['borrowed_by']} na projekt {tool_dict['project_name']}")
            else:
                tool_dict['borrowed_by'] = 'Brak danych'
                tool_dict['project_name'] = 'Brak danych'
                tool_dict['borrowed_date'] = ''
        else:
            tool_dict['borrowed_by'] = None
            tool_dict['project_name'] = None
            tool_dict['borrowed_date'] = None
        
        tools_data.append(tool_dict)
    
    return jsonify(tools_data), 200

@app.route('/api/tools', methods=['POST'])
# @jwt_required()
def add_tool():
    """Dodaje nowe narzędzie"""
    data = request.get_json()
    
    tool = Tool(
        name=data['name'],
        category=data.get('category', 'Inne'),
        description=data.get('description', '')
    )
    tool.generate_code()
    
    db.session.add(tool)
    db.session.commit()
    
    print(f"✓ Dodano narzędzie: {tool.name} (kod: {tool.code})")
    return jsonify(tool.to_dict()), 201

@app.route('/api/tools/<int:tool_id>/borrow', methods=['POST'])
# @jwt_required()
def borrow_tool(tool_id):
    """Wypożycza narzędzie"""
    tool = Tool.query.get_or_404(tool_id)
    data = request.get_json()
    user_id = data.get('user_id')
    
    if not user_id:
        admin = User.query.filter_by(username='admin').first()
        user_id = admin.id if admin else 1
    if tool.status != 'available':
        return jsonify({'message': 'Narzędzie nie jest dostępne'}), 400
    # Debug
    print(f"Wypożyczanie: tool_id={tool_id}, user_id={user_id}, project_id={data.get('project_id')}")
    
    # Utwórz wpis w historii
    history = ToolHistory(
        tool_id=tool_id,
        user_id=user_id,
        project_id=data.get('project_id'),
        action='borrowed',
        notes=data.get('notes', '')
    )
    
    # Zmień status narzędzia
    tool.status = 'borrowed'
    
    db.session.add(history)
    db.session.commit()
    
    print(f"✓ Wypożyczono: {tool.name} przez user_id={user_id}")
    return jsonify({'message': 'Narzędzie wypożyczone pomyślnie'}), 200

@app.route('/api/debug/history')
def debug_history():
    """Debug - pokazuje historię wypożyczeń"""
    history = ToolHistory.query.filter_by(returned_at=None).all()
    result = []
    for h in history:
        result.append({
            'tool': h.tool.name if h.tool else 'Brak',
            'tool_id': h.tool_id,
            'user': h.user.username if h.user else 'Brak',
            'user_id': h.user_id,
            'project': h.project.name if h.project else 'Brak',
            'project_id': h.project_id,
            'borrowed_at': h.borrowed_at.isoformat() if h.borrowed_at else None
        })
    return jsonify(result)

@app.route('/api/tools/<int:tool_id>/return', methods=['POST'])
# @jwt_required()
def return_tool(tool_id):
    """Zwraca narzędzie"""
    tool = Tool.query.get_or_404(tool_id)
    data = request.get_json()
    user_id = data.get('user_id', 1)
    
    # Znajdź ostatnie wypożyczenie
    history = ToolHistory.query.filter_by(
        tool_id=tool_id,
        user_id=user_id,
        returned_at=None
    ).first()
    
    if not history:
        return jsonify({'message': 'Nie znaleziono aktywnego wypożyczenia'}), 400
    
    # Zaktualizuj historię
    history.returned_at = datetime.utcnow()
    history.notes = data.get('notes', history.notes)
    
    # Jeśli to wymuszony zwrot, dodaj info
    if 'wymuszony' in notes.lower():
        history.notes = f"{notes} (przez user_id: {user_id})"
    else:
        history.notes = notes
    
    # Zmień status narzędzia
    tool.status = 'available'
    
    db.session.commit()
    
    print(f"✓ Zwrócono: {tool.name}")
    return jsonify({'message': 'Narzędzie zwrócone pomyślnie'}), 200

@app.route('/api/tools/<int:tool_id>/history', methods=['GET'])
# @jwt_required()
def get_tool_history(tool_id):
    """Pobiera historię narzędzia"""
    history = ToolHistory.query.filter_by(tool_id=tool_id).order_by(ToolHistory.borrowed_at.desc()).all()
    return jsonify([h.to_dict() for h in history]), 200

@app.route('/api/stats')
def get_stats():
    """Zwraca statystyki systemu"""
    try:
        total_tools = Tool.query.count()
        available_tools = Tool.query.filter_by(status='available').count()
        borrowed_tools = Tool.query.filter_by(status='borrowed').count()
        total_projects = Project.query.count()
        total_history = ToolHistory.query.count()
        total_users = User.query.count()
        
        # Najpopularniejsze narzędzie
        from sqlalchemy import func
        popular = db.session.query(
            Tool.name,
            func.count(ToolHistory.id).label('count')
        ).join(ToolHistory, Tool.id == ToolHistory.tool_id).group_by(
            Tool.id
        ).order_by(
            func.count(ToolHistory.id).desc()
        ).first()
        
        # Najaktywniejszy użytkownik
        active_user = db.session.query(
            User.username,
            func.count(ToolHistory.id).label('count')
        ).join(ToolHistory, User.id == ToolHistory.user_id).group_by(
            User.id
        ).order_by(
            func.count(ToolHistory.id).desc()
        ).first()
        
        return jsonify({
            'total_tools': total_tools,
            'available_tools': available_tools,
            'borrowed_tools': borrowed_tools,
            'total_projects': total_projects,
            'total_rentals': total_history,
            'total_users': total_users,
            'most_popular_tool': popular[0] if popular else 'Brak danych',
            'most_popular_count': popular[1] if popular else 0,
            'most_active_user': active_user[0] if active_user else 'Brak danych',
            'most_active_user_count': active_user[1] if active_user else 0
        }), 200
    except Exception as e:
        print(f"Błąd w statystykach: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/tools/<int:tool_id>', methods=['PUT'])
def update_tool(tool_id):
    """Aktualizuje dane narzędzia"""
    tool = Tool.query.get_or_404(tool_id)
    data = request.get_json()
    
    # Sprawdź unikalność kodu jeśli się zmienił
    new_code = data.get('code')
    if new_code and new_code != tool.code:
        existing = Tool.query.filter_by(code=new_code).first()
        if existing:
            return jsonify({'message': 'Kod już istnieje!'}), 400
        tool.code = new_code
    
    # Aktualizuj pozostałe pola
    tool.name = data.get('name', tool.name)
    tool.category = data.get('category', tool.category)
    tool.description = data.get('description', tool.description)
    
    db.session.commit()
    
    print(f"✓ Zaktualizowano narzędzie: {tool.name}")
    return jsonify(tool.to_dict()), 200

# === PROJEKTY ===
@app.route('/api/projects', methods=['GET'])
# @jwt_required()
def get_projects():
    """Pobiera listę projektów"""
    projects = Project.query.all()
    return jsonify([{
        'id': p.id,
        'name': p.name,
        'description': p.description,
        'status': p.status
    } for p in projects]), 200

@app.route('/api/projects', methods=['POST'])
# @jwt_required()
def add_project():
    """Dodaje nowy projekt"""
    data = request.get_json()
    
    project = Project(
        name=data['name'],
        description=data.get('description', '')
    )
    
    db.session.add(project)
    db.session.commit()
    
    print(f"✓ Dodano projekt: {project.name}")
    return jsonify({
        'id': project.id,
        'name': project.name,
        'description': project.description
    }), 201
@app.route('/api/export/excel')
def export_to_excel():
    """Eksportuje dane do pliku Excel"""
    # Utwórz nowy plik Excel
    wb = Workbook()
    
    # Arkusz z narzędziami
    ws_tools = wb.active
    ws_tools.title = "Narzędzia"
    ws_tools.append(["ID", "Kod", "Nazwa", "Kategoria", "Status", "Opis"])
    
    tools = Tool.query.all()
    for tool in tools:
        ws_tools.append([
            tool.id,
            tool.code,
            tool.name,
            tool.category,
            tool.status,
            tool.description
        ])
    
    # Arkusz z historią
    ws_history = wb.create_sheet("Historia")
    ws_history.append(["ID", "Narzędzie", "Użytkownik", "Projekt", "Akcja", "Data wypożyczenia", "Data zwrotu"])
    
    history = ToolHistory.query.all()
    for h in history:
        ws_history.append([
            h.id,
            h.tool.name if h.tool else "",
            h.user.username if h.user else "",
            h.project.name if h.project else "",
            h.action,
            h.borrowed_at.strftime("%Y-%m-%d %H:%M") if h.borrowed_at else "",
            h.returned_at.strftime("%Y-%m-%d %H:%M") if h.returned_at else "Nie zwrócono"
        ])
    
    # Arkusz z projektami
    ws_projects = wb.create_sheet("Projekty")
    ws_projects.append(["ID", "Nazwa", "Opis", "Status"])
    
    projects = Project.query.all()
    for project in projects:
        ws_projects.append([
            project.id,
            project.name,
            project.description,
            project.status
        ])
    
    # Zapisz do pliku tymczasowego
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=f'narzedzia_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/tools/<int:tool_id>', methods=['DELETE'])
def delete_tool(tool_id):
    """Usuwa narzędzie"""
    tool = Tool.query.get_or_404(tool_id)
    
    # Sprawdź czy narzędzie nie jest wypożyczone
    if tool.status == 'borrowed':
        return jsonify({'message': 'Nie można usunąć wypożyczonego narzędzia!'}), 400
    
    # Usuń powiązaną historię
    ToolHistory.query.filter_by(tool_id=tool_id).delete()
    
    # Usuń narzędzie
    db.session.delete(tool)
    db.session.commit()
    
    print(f"✓ Usunięto narzędzie: {tool.name}")
    return jsonify({'message': 'Narzędzie usunięte'}), 200

@app.route('/api/users', methods=['GET'])
def get_users():
    """Pobiera listę użytkowników"""
    users = User.query.all()
    return jsonify([{
        'id': u.id,
        'username': u.username,
        'email': u.email,
        'role': u.role,
        'created_at': u.created_at.isoformat() if u.created_at else None
    } for u in users]), 200

@app.route('/api/users', methods=['POST'])
def add_user():
    """Dodaje nowego użytkownika"""
    data = request.get_json()
    
    # Sprawdź czy użytkownik już istnieje
    if User.query.filter_by(username=data['username']).first():
        return jsonify({'message': 'Użytkownik o tej nazwie już istnieje!'}), 400
    
    if User.query.filter_by(email=data['email']).first():
        return jsonify({'message': 'Ten email jest już używany!'}), 400
    
    user = User(
        username=data['username'],
        email=data['email'],
        password_hash=generate_password_hash(data['password']),
        role=data.get('role', 'worker')
    )
    
    db.session.add(user)
    db.session.commit()
    
    print(f"✓ Dodano użytkownika: {user.username}")
    return jsonify({'message': 'Użytkownik dodany pomyślnie', 'id': user.id}), 201

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    """Usuwa użytkownika"""
    user = User.query.get_or_404(user_id)
    
    # Nie pozwól usunąć admina
    if user.username == 'admin':
        return jsonify({'message': 'Nie można usunąć głównego administratora!'}), 400
    
    # Sprawdź czy użytkownik ma wypożyczone narzędzia
    active_rentals = ToolHistory.query.filter_by(
        user_id=user_id,
        returned_at=None
    ).count()
    
    if active_rentals > 0:
        return jsonify({'message': f'Użytkownik ma {active_rentals} niewróconych narzędzi!'}), 400
    
    db.session.delete(user)
    db.session.commit()
    
    print(f"✓ Usunięto użytkownika: {user.username}")
    return jsonify({'message': 'Użytkownik usunięty'}), 200

@app.route('/api/users/<int:user_id>/reset-password', methods=['POST'])
def reset_password(user_id):
    """Resetuje hasło użytkownika"""
    user = User.query.get_or_404(user_id)
    data = request.get_json()
    
    new_password = data.get('password', 'haslo123')
    user.password_hash = generate_password_hash(new_password)
    
    db.session.commit()
    
    print(f"✓ Zresetowano hasło dla: {user.username}")
    return jsonify({'message': f'Hasło zresetowane. Nowe hasło: {new_password}'}), 200
# === URUCHOMIENIE SERWERA ===
if __name__ == '__main__':
    print("\n" + "="*50)
    print("SYSTEM ZARZĄDZANIA NARZĘDZIAMI")
    print("="*50)
    print("\n✓ Serwer uruchomiony!")
    print("✓ Adres: http://localhost:5000")
    print("\n✓ Dane logowania:")
    print("  Użytkownik: admin")
    print("  Hasło: admin123")
    print("\n" + "="*50 + "\n")
    
    app.run(debug=True, host='0.0.0.0', port=5000)

@app.route('/api/tools/<int:tool_id>/qr')
def get_tool_qr(tool_id):
    """Generuje kod QR dla narzędzia"""
    tool = Tool.query.get_or_404(tool_id)
    qr_base64 = generate_qr_code(tool.code, tool.name)
    
    return jsonify({
        'tool_name': tool.name,
        'tool_code': tool.code,
        'qr_image': qr_base64
    })
@app.route('/api/tools/search')
def search_tools():
    """Wyszukuje narzędzia"""
    query = request.args.get('q', '')
    category = request.args.get('category', '')
    status = request.args.get('status', '')
    
    tools_query = Tool.query
    
    if query:
        tools_query = tools_query.filter(
            db.or_(
                Tool.name.contains(query),
                Tool.code.contains(query),
                Tool.description.contains(query)
            )
        )
    
    if category:
        tools_query = tools_query.filter_by(category=category)
    
    if status:
        tools_query = tools_query.filter_by(status=status)
    
    tools = tools_query.all()
    return jsonify([tool.to_dict() for tool in tools])